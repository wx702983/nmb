
'''

Python的openpyxl的操作


接口自动化测试流程
1、接口文档---接口测试用例---Excel表里
2、Excel表里读取测试用例---得到数据
3、执行测试---requests去执行---执行结果
4、执行结果---测试结果---对比--是否通过--最终结果
5、最终结果写入Excel里---pass---failed

封装函数步骤：
1、实现功能的代码写出来
2、参数化---可变参数---形参
3、返回值---函数需要给别人使用的东西


读取数据：第三方库 openpyxl
Excel表格操作：
1、读取
2、写入、写入生效一定要保存文件、保存之前关闭文件
Excel表单三大对象
1、文档---加载进来---工作簿对象---load_workbook---最好放在同级路径
2、表单对象
3、单元格对象


'''


# 注册接口请求
import requests

url_api_reg = "http://8.129.91.152:8766/futureloan/member/register"
data_api_reg = {
    "mobile_phone": "15549083142",
    "pwd": "12345678",
    "type": "1",
    "reg_name": "mengmeng"
}

head_reg = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
response = requests.post(url=url_api_reg, json=data_api_reg, headers=head_reg)
print(response.json())















# 发送接口请求函数
import requests,pprint,jsonpath
def func_api(url_api,data_api):
    head_api = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    response = requests.post(url=url_api, json=data_api, headers=head_api)
    print(response.json())
    return response.json()



#导入指定部分
from openpyxl import load_workbook
#加载工作簿
wb = load_workbook("test_case_api.xlsx")
#获取表单
sheet = wb["register"]
#获取单元格属性
cell = sheet.cell(row=1,column=1).value
#对单元格赋值--写入
cell ="用例编号"
print(cell)
wb.save("test_case_api.xlsx")




#Excel表里读取测试用例---得到数据
from openpyxl import load_workbook
wb = load_workbook("test_case_api.xlsx")
sheet = wb["register"]
#自动获取最大行号、列号
cases = [] #定义空列表
max_row = sheet.max_row
for i in range(2,max_row+1):  #取头不取尾  +1
    case = dict(
    case_id = sheet.cell(row=i,column=1).value,
    url = sheet.cell(row=i,column=5).value,
    data = sheet.cell(row=i,column=6).value,
    expected_result = sheet.cell(row=i,column=7).value)
    cases.append(case)
print(cases)



#Excel表里读取测试用例---得到数据---函数参数化

from openpyxl import load_workbook
def read_data(filename,sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    #自动获取最大行号、列号
    cases = [] #定义空列表
    max_row = sheet.max_row
    for i in range(2,max_row+1):  #取头不取尾  +1
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expected_result = sheet.cell(row=i,column=7).value)
        cases.append(case)
    return cases
result = read_data("test_case_api.xlsx","login")
print(result)




#对单元格赋值--写入

from openpyxl import load_workbook
wb = load_workbook("test_case_api.xlsx")
sheet = wb["register"]
cell = sheet.cell(row=2,column=18).value="Passed"
wb.save("test_case_api.xlsx")



#写入结果函数

from openpyxl import load_workbook
def write_data(filename,sheetname,final_result,row,column):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    cell = sheet.cell(row=row,column=row).value=final_result
    wb.save(filename)















# 登录接口函数
import pprint
def func_log(url_):
    url_api_log = "http://8.129.91.152:8766/futureloan/member/login"
    data_api_log = {
        "mobile_phone": "15815541706",
        "pwd": "12345678",
    }
    head_log = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    response = requests.post(url=url_api_log, json=data_api_log, headers=head_log)
    pprint.pprint(response.json())
    res = response.json()

